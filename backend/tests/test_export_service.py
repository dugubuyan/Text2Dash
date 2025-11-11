"""
导出服务测试
"""
import pytest
import asyncio
import sys
from pathlib import Path
from datetime import datetime
from io import BytesIO

# 添加项目根目录到Python路径
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from backend.services.export_service import ExportService, ReportData
from backend.services.dto import DataMetadata


@pytest.fixture
def export_service():
    """创建导出服务实例"""
    return ExportService()


@pytest.fixture
def sample_report_data():
    """创建示例报表数据"""
    data = [
        {"name": "Alice", "age": 25, "score": 95.5, "city": "New York"},
        {"name": "Bob", "age": 30, "score": 88.0, "city": "Los Angeles"},
        {"name": "Charlie", "age": 35, "score": 92.3, "city": "Chicago"},
        {"name": "David", "age": 28, "score": 87.5, "city": "Houston"},
        {"name": "Eve", "age": 32, "score": 91.0, "city": "Phoenix"},
    ]
    
    metadata = DataMetadata(
        columns=["name", "age", "score", "city"],
        column_types={
            "name": "TEXT",
            "age": "INTEGER",
            "score": "REAL",
            "city": "TEXT"
        },
        row_count=5
    )
    
    chart_config = {
        "type": "bar",
        "title": {"text": "Student Scores"},
        "xAxis": {"type": "category", "data": ["Alice", "Bob", "Charlie", "David", "Eve"]},
        "yAxis": {"type": "value"},
        "series": [{"data": [95.5, 88.0, 92.3, 87.5, 91.0], "type": "bar"}]
    }
    
    return ReportData(
        title="Student Performance Report",
        summary="This report shows the performance of 5 students. The average score is 90.86. Alice has the highest score of 95.5.",
        data=data,
        metadata=metadata,
        chart_config=chart_config,
        sql_query="SELECT name, age, score, city FROM students WHERE score > 85"
    )


@pytest.mark.asyncio
async def test_export_to_pdf_basic(export_service, sample_report_data):
    """测试基本PDF导出功能"""
    pdf_bytes = await export_service.export_to_pdf(sample_report_data)
    
    # 验证返回的是字节数据
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    
    # 验证PDF文件头
    assert pdf_bytes.startswith(b'%PDF')
    
    print(f"✓ PDF生成成功: {len(pdf_bytes)} bytes")


@pytest.mark.asyncio
async def test_export_to_pdf_without_chart(export_service, sample_report_data):
    """测试不包含图表的PDF导出"""
    pdf_bytes = await export_service.export_to_pdf(
        sample_report_data,
        include_chart=False
    )
    
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    assert pdf_bytes.startswith(b'%PDF')
    
    print(f"✓ PDF生成成功（无图表）: {len(pdf_bytes)} bytes")


@pytest.mark.asyncio
async def test_export_to_pdf_without_data_table(export_service, sample_report_data):
    """测试不包含数据表格的PDF导出"""
    pdf_bytes = await export_service.export_to_pdf(
        sample_report_data,
        include_data_table=False
    )
    
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    assert pdf_bytes.startswith(b'%PDF')
    
    print(f"✓ PDF生成成功（无数据表格）: {len(pdf_bytes)} bytes")


@pytest.mark.asyncio
async def test_export_to_pdf_with_max_rows(export_service, sample_report_data):
    """测试限制行数的PDF导出"""
    pdf_bytes = await export_service.export_to_pdf(
        sample_report_data,
        max_rows=3
    )
    
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    
    print(f"✓ PDF生成成功（限制3行）: {len(pdf_bytes)} bytes")


@pytest.mark.asyncio
async def test_export_to_excel_basic(export_service, sample_report_data):
    """测试基本Excel导出功能"""
    excel_bytes = await export_service.export_to_excel(sample_report_data)
    
    # 验证返回的是字节数据
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    # 验证Excel文件头（ZIP格式）
    assert excel_bytes.startswith(b'PK')
    
    print(f"✓ Excel生成成功: {len(excel_bytes)} bytes")


@pytest.mark.asyncio
async def test_export_to_excel_verify_content(export_service, sample_report_data):
    """测试Excel内容验证"""
    from openpyxl import load_workbook
    
    excel_bytes = await export_service.export_to_excel(sample_report_data)
    
    # 加载Excel文件
    wb = load_workbook(BytesIO(excel_bytes))
    
    # 验证工作表
    assert "Data" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Metadata" in wb.sheetnames
    
    # 验证数据工作表
    ws_data = wb["Data"]
    
    # 验证表头
    headers = [cell.value for cell in ws_data[1]]
    assert headers == ["name", "age", "score", "city"]
    
    # 验证数据行数（表头 + 数据）
    assert ws_data.max_row == 6  # 1 header + 5 data rows
    
    # 验证第一行数据
    first_row = [cell.value for cell in ws_data[2]]
    assert first_row == ["Alice", 25, 95.5, "New York"]
    
    print(f"✓ Excel内容验证成功: {len(wb.sheetnames)} 个工作表")


@pytest.mark.asyncio
async def test_export_to_excel_without_metadata(export_service, sample_report_data):
    """测试不包含元信息的Excel导出"""
    from openpyxl import load_workbook
    
    excel_bytes = await export_service.export_to_excel(
        sample_report_data,
        include_metadata=False
    )
    
    wb = load_workbook(BytesIO(excel_bytes))
    
    assert "Data" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Metadata" not in wb.sheetnames
    
    print(f"✓ Excel生成成功（无元信息）: {len(wb.sheetnames)} 个工作表")


@pytest.mark.asyncio
async def test_export_to_excel_without_summary(export_service, sample_report_data):
    """测试不包含总结的Excel导出"""
    from openpyxl import load_workbook
    
    excel_bytes = await export_service.export_to_excel(
        sample_report_data,
        include_summary=False
    )
    
    wb = load_workbook(BytesIO(excel_bytes))
    
    assert "Data" in wb.sheetnames
    assert "Summary" not in wb.sheetnames
    assert "Metadata" in wb.sheetnames
    
    print(f"✓ Excel生成成功（无总结）: {len(wb.sheetnames)} 个工作表")


@pytest.mark.asyncio
async def test_export_large_dataset(export_service):
    """测试大数据集导出"""
    # 创建1000行数据
    large_data = [
        {
            "id": i,
            "name": f"User_{i}",
            "value": i * 10.5,
            "category": f"Category_{i % 10}"
        }
        for i in range(1000)
    ]
    
    metadata = DataMetadata(
        columns=["id", "name", "value", "category"],
        column_types={
            "id": "INTEGER",
            "name": "TEXT",
            "value": "REAL",
            "category": "TEXT"
        },
        row_count=1000
    )
    
    report_data = ReportData(
        title="Large Dataset Report",
        summary="This report contains 1000 rows of data for performance testing.",
        data=large_data,
        metadata=metadata
    )
    
    # 测试PDF导出（限制100行）
    pdf_bytes = await export_service.export_to_pdf(report_data, max_rows=100)
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    
    # 测试Excel导出（全部数据）
    excel_bytes = await export_service.export_to_excel(report_data)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    # 验证Excel包含所有数据
    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(excel_bytes))
    ws_data = wb["Data"]
    assert ws_data.max_row == 1001  # 1 header + 1000 data rows
    
    print(f"✓ 大数据集导出成功: PDF={len(pdf_bytes)} bytes, Excel={len(excel_bytes)} bytes")


@pytest.mark.asyncio
async def test_export_empty_data(export_service):
    """测试空数据导出"""
    empty_data = []
    
    metadata = DataMetadata(
        columns=["col1", "col2"],
        column_types={"col1": "TEXT", "col2": "INTEGER"},
        row_count=0
    )
    
    report_data = ReportData(
        title="Empty Report",
        summary="This report has no data.",
        data=empty_data,
        metadata=metadata
    )
    
    # PDF导出应该成功（只有标题和总结）
    pdf_bytes = await export_service.export_to_pdf(report_data)
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    
    # Excel导出应该成功（只有表头）
    excel_bytes = await export_service.export_to_excel(report_data)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    print(f"✓ 空数据导出成功")


@pytest.mark.asyncio
async def test_export_special_characters(export_service):
    """测试特殊字符处理"""
    data = [
        {"name": "Test & Co.", "value": 100, "note": "Price: $50 < $100"},
        {"name": "Company \"A\"", "value": 200, "note": "Discount: 10% > 5%"},
        {"name": "Café", "value": 150, "note": "Location: São Paulo"},
    ]
    
    metadata = DataMetadata(
        columns=["name", "value", "note"],
        column_types={"name": "TEXT", "value": "INTEGER", "note": "TEXT"},
        row_count=3
    )
    
    report_data = ReportData(
        title="Special Characters Test",
        summary="Testing special characters: & < > \" ' % $ €",
        data=data,
        metadata=metadata
    )
    
    # 测试PDF导出
    pdf_bytes = await export_service.export_to_pdf(report_data)
    assert isinstance(pdf_bytes, bytes)
    assert len(pdf_bytes) > 0
    
    # 测试Excel导出
    excel_bytes = await export_service.export_to_excel(report_data)
    assert isinstance(excel_bytes, bytes)
    assert len(excel_bytes) > 0
    
    print(f"✓ 特殊字符处理成功")


if __name__ == "__main__":
    # 运行测试
    pytest.main([__file__, "-v", "-s"])
