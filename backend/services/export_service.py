"""
导出服务
提供PDF和Excel格式的报表导出功能
"""
import io
import json
from datetime import datetime
from typing import List, Dict, Any, Optional
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.enums import TA_CENTER, TA_LEFT

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .dto import DataMetadata
from ..utils.logger import get_logger

logger = get_logger(__name__)


class ReportData:
    """报表数据封装类"""
    def __init__(
        self,
        title: str,
        summary: str,
        data: List[Dict[str, Any]],
        metadata: DataMetadata,
        chart_config: Optional[Dict[str, Any]] = None,
        chart_image: Optional[bytes] = None,
        sql_query: Optional[str] = None
    ):
        self.title = title
        self.summary = summary
        self.data = data
        self.metadata = metadata
        self.chart_config = chart_config
        self.chart_image = chart_image
        self.sql_query = sql_query
        self.generated_at = datetime.now()


class ExportService:
    """导出服务类"""
    
    def __init__(self):
        """初始化导出服务"""
        logger.info("导出服务初始化完成")
    
    async def export_to_pdf(
        self,
        report_data: ReportData,
        include_chart: bool = True,
        include_data_table: bool = True,
        max_rows: int = 100
    ) -> bytes:
        """
        生成PDF文件
        
        包含：
        - 标题和生成时间
        - 报告总结
        - 图表图片（如果提供）
        - 数据表格
        - 数据来源信息
        
        Args:
            report_data: 报表数据对象
            include_chart: 是否包含图表
            include_data_table: 是否包含数据表格
            max_rows: 数据表格最大行数（避免PDF过大）
        
        Returns:
            PDF文件的字节数据
        
        Raises:
            Exception: 如果PDF生成失败
        """
        try:
            logger.info(
                f"开始生成PDF: title='{report_data.title}', "
                f"data_rows={len(report_data.data)}, "
                f"include_chart={include_chart}, "
                f"include_data_table={include_data_table}"
            )
            
            # 注册中文字体（使用系统自带的字体）
            # macOS 使用 PingFang SC，Windows 使用 SimSun，Linux 使用 Noto Sans CJK
            try:
                import platform
                system = platform.system()
                
                if system == 'Darwin':  # macOS
                    # 尝试使用 PingFang SC
                    try:
                        pdfmetrics.registerFont(TTFont('Chinese', '/System/Library/Fonts/PingFang.ttc', subfontIndex=0))
                        chinese_font = 'Chinese'
                    except:
                        # 备选：使用 STHeiti
                        try:
                            pdfmetrics.registerFont(TTFont('Chinese', '/System/Library/Fonts/STHeiti Light.ttc'))
                            chinese_font = 'Chinese'
                        except:
                            logger.warning("无法加载中文字体，使用默认字体")
                            chinese_font = 'Helvetica'
                elif system == 'Windows':
                    # Windows 使用 SimSun
                    try:
                        pdfmetrics.registerFont(TTFont('Chinese', 'C:\\Windows\\Fonts\\simsun.ttc'))
                        chinese_font = 'Chinese'
                    except:
                        logger.warning("无法加载中文字体，使用默认字体")
                        chinese_font = 'Helvetica'
                else:  # Linux
                    # 尝试使用 Noto Sans CJK
                    try:
                        pdfmetrics.registerFont(TTFont('Chinese', '/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc'))
                        chinese_font = 'Chinese'
                    except:
                        logger.warning("无法加载中文字体，使用默认字体")
                        chinese_font = 'Helvetica'
            except Exception as e:
                logger.warning(f"字体加载失败: {e}，使用默认字体")
                chinese_font = 'Helvetica'
            
            # 创建PDF文档
            buffer = BytesIO()
            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                rightMargin=0.75*inch,
                leftMargin=0.75*inch,
                topMargin=1*inch,
                bottomMargin=0.75*inch
            )
            
            # 构建PDF内容
            story = []
            styles = getSampleStyleSheet()
            
            # 添加标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontName=chinese_font,
                fontSize=18,
                textColor=colors.HexColor('#1f2937'),
                spaceAfter=12,
                alignment=TA_CENTER
            )
            story.append(Paragraph(report_data.title, title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # 添加生成时间
            time_text = f"Generated: {report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"
            time_style = ParagraphStyle(
                'TimeStyle',
                parent=styles['Normal'],
                fontName=chinese_font,
                fontSize=10,
                textColor=colors.HexColor('#6b7280'),
                alignment=TA_CENTER
            )
            story.append(Paragraph(time_text, time_style))
            story.append(Spacer(1, 0.3*inch))
            
            # 添加报告总结
            if report_data.summary:
                summary_title_style = ParagraphStyle(
                    'SummaryTitle',
                    parent=styles['Heading2'],
                    fontName=chinese_font,
                    fontSize=14,
                    textColor=colors.HexColor('#374151'),
                    spaceAfter=8
                )
                story.append(Paragraph("Summary", summary_title_style))
                
                summary_style = ParagraphStyle(
                    'SummaryText',
                    parent=styles['Normal'],
                    fontName=chinese_font,
                    fontSize=11,
                    textColor=colors.HexColor('#1f2937'),
                    spaceAfter=12,
                    leading=16
                )
                # 处理换行符
                summary_paragraphs = report_data.summary.split('\n')
                for para in summary_paragraphs:
                    if para.strip():
                        story.append(Paragraph(para, summary_style))
                
                story.append(Spacer(1, 0.3*inch))
            
            # 添加图表图片
            if include_chart and report_data.chart_image:
                try:
                    chart_title_style = ParagraphStyle(
                        'ChartTitle',
                        parent=styles['Heading2'],
                        fontName=chinese_font,
                        fontSize=14,
                        textColor=colors.HexColor('#374151'),
                        spaceAfter=8
                    )
                    story.append(Paragraph("Chart", chart_title_style))
                    
                    # 将图片字节转换为Image对象
                    img = Image(BytesIO(report_data.chart_image))
                    
                    # 调整图片大小以适应页面
                    available_width = doc.width
                    available_height = doc.height * 0.4  # 使用40%的页面高度
                    
                    # 计算缩放比例
                    img_width, img_height = img.imageWidth, img.imageHeight
                    width_ratio = available_width / img_width
                    height_ratio = available_height / img_height
                    scale_ratio = min(width_ratio, height_ratio, 1.0)  # 不放大
                    
                    img.drawWidth = img_width * scale_ratio
                    img.drawHeight = img_height * scale_ratio
                    
                    story.append(img)
                    story.append(Spacer(1, 0.3*inch))
                    
                    logger.debug(f"图表已添加到PDF: size={img.drawWidth}x{img.drawHeight}")
                except Exception as e:
                    logger.warning(f"添加图表到PDF失败: {e}")
            
            # 添加数据表格
            if include_data_table and report_data.data:
                data_title_style = ParagraphStyle(
                    'DataTitle',
                    parent=styles['Heading2'],
                    fontName=chinese_font,
                    fontSize=14,
                    textColor=colors.HexColor('#374151'),
                    spaceAfter=8
                )
                story.append(Paragraph("Data", data_title_style))
                
                # 限制行数
                display_data = report_data.data[:max_rows]
                
                # 构建表格数据
                table_data = []
                
                # 表头
                headers = report_data.metadata.columns
                table_data.append(headers)
                
                # 数据行
                for row in display_data:
                    table_row = []
                    for col in headers:
                        value = row.get(col, '')
                        # 转换为字符串并限制长度
                        str_value = str(value) if value is not None else ''
                        if len(str_value) > 50:
                            str_value = str_value[:47] + '...'
                        table_row.append(str_value)
                    table_data.append(table_row)
                
                # 创建表格
                # 计算列宽
                num_cols = len(headers)
                col_width = doc.width / num_cols
                
                table = Table(table_data, colWidths=[col_width] * num_cols)
                
                # 设置表格样式
                table.setStyle(TableStyle([
                    # 表头样式
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3b82f6')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), chinese_font),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    
                    # 数据行样式
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#1f2937')),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 1), (-1, -1), chinese_font),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                    
                    # 网格线
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e5e7eb')),
                    
                    # 交替行背景色
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), 
                     [colors.white, colors.HexColor('#f9fafb')])
                ]))
                
                story.append(table)
                
                # 如果数据被截断，添加说明
                if len(report_data.data) > max_rows:
                    note_style = ParagraphStyle(
                        'NoteStyle',
                        parent=styles['Normal'],
                        fontName=chinese_font,
                        fontSize=9,
                        textColor=colors.HexColor('#6b7280'),
                        spaceAfter=8
                    )
                    note_text = f"Note: Showing {max_rows} of {len(report_data.data)} rows"
                    story.append(Spacer(1, 0.1*inch))
                    story.append(Paragraph(note_text, note_style))
                
                logger.debug(
                    f"数据表格已添加到PDF: rows={len(display_data)}, "
                    f"cols={len(headers)}"
                )
            
            # 添加元信息
            story.append(Spacer(1, 0.3*inch))
            meta_title_style = ParagraphStyle(
                'MetaTitle',
                parent=styles['Heading2'],
                fontName=chinese_font,
                fontSize=14,
                textColor=colors.HexColor('#374151'),
                spaceAfter=8
            )
            story.append(Paragraph("Metadata", meta_title_style))
            
            meta_style = ParagraphStyle(
                'MetaText',
                parent=styles['Normal'],
                fontName=chinese_font,
                fontSize=10,
                textColor=colors.HexColor('#4b5563')
            )
            
            meta_info = [
                f"Total Rows: {report_data.metadata.row_count}",
                f"Total Columns: {len(report_data.metadata.columns)}",
                f"Columns: {', '.join(report_data.metadata.columns)}"
            ]
            
            for info in meta_info:
                story.append(Paragraph(info, meta_style))
                story.append(Spacer(1, 0.05*inch))
            
            # 生成PDF
            doc.build(story)
            
            # 获取PDF字节数据
            pdf_bytes = buffer.getvalue()
            buffer.close()
            
            logger.info(
                f"PDF生成完成: size={len(pdf_bytes)} bytes, "
                f"pages_estimated={(len(report_data.data) // 20) + 1}"
            )
            
            return pdf_bytes
            
        except Exception as e:
            logger.error(
                f"PDF生成失败",
                extra={
                    "title": report_data.title,
                    "data_rows": len(report_data.data),
                    "error": str(e)
                },
                exc_info=True
            )
            raise Exception(f"PDF生成失败: {str(e)}")
    
    async def export_to_excel(
        self,
        report_data: ReportData,
        include_metadata: bool = True,
        include_summary: bool = True
    ) -> bytes:
        """
        生成Excel文件
        
        包含：
        - 数据工作表（原始数据）
        - 元信息工作表（列信息、数据类型等）
        - 总结工作表（报告总结和生成信息）
        
        Args:
            report_data: 报表数据对象
            include_metadata: 是否包含元信息工作表
            include_summary: 是否包含总结工作表
        
        Returns:
            Excel文件的字节数据
        
        Raises:
            Exception: 如果Excel生成失败
        """
        try:
            logger.info(
                f"开始生成Excel: title='{report_data.title}', "
                f"data_rows={len(report_data.data)}, "
                f"include_metadata={include_metadata}, "
                f"include_summary={include_summary}"
            )
            
            # 创建工作簿
            wb = Workbook()
            
            # 移除默认工作表
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # 定义样式
            header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='3B82F6', end_color='3B82F6', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            data_font = Font(name='Arial', size=10)
            data_alignment = Alignment(horizontal='left', vertical='center')
            
            border = Border(
                left=Side(style='thin', color='E5E7EB'),
                right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'),
                bottom=Side(style='thin', color='E5E7EB')
            )
            
            # 工作表1: 数据
            ws_data = wb.create_sheet("Data", 0)
            
            # 写入表头
            headers = report_data.metadata.columns
            for col_idx, header in enumerate(headers, 1):
                cell = ws_data.cell(row=1, column=col_idx, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            
            # 写入数据
            for row_idx, row_data in enumerate(report_data.data, 2):
                for col_idx, col_name in enumerate(headers, 1):
                    value = row_data.get(col_name, '')
                    cell = ws_data.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = border
                    
                    # 交替行背景色
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(
                            start_color='F9FAFB',
                            end_color='F9FAFB',
                            fill_type='solid'
                        )
            
            # 自动调整列宽
            for col_idx, col_name in enumerate(headers, 1):
                # 计算列宽（基于列名和数据）
                max_length = len(str(col_name))
                for row_data in report_data.data[:100]:  # 只检查前100行
                    value = row_data.get(col_name, '')
                    max_length = max(max_length, len(str(value)))
                
                # 设置列宽（限制最大宽度）
                adjusted_width = min(max_length + 2, 50)
                ws_data.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
            
            # 冻结首行
            ws_data.freeze_panes = 'A2'
            
            logger.debug(
                f"数据工作表已创建: rows={len(report_data.data)}, "
                f"cols={len(headers)}"
            )
            
            # 工作表2: 总结
            if include_summary:
                ws_summary = wb.create_sheet("Summary", 1)
                
                # 标题
                title_cell = ws_summary.cell(row=1, column=1, value="Report Summary")
                title_cell.font = Font(name='Arial', size=14, bold=True, color='1F2937')
                ws_summary.merge_cells('A1:B1')
                
                # 报表信息
                info_data = [
                    ("Title", report_data.title),
                    ("Generated At", report_data.generated_at.strftime('%Y-%m-%d %H:%M:%S')),
                    ("Total Rows", report_data.metadata.row_count),
                    ("Total Columns", len(report_data.metadata.columns)),
                ]
                
                for row_idx, (label, value) in enumerate(info_data, 3):
                    label_cell = ws_summary.cell(row=row_idx, column=1, value=label)
                    label_cell.font = Font(name='Arial', size=10, bold=True)
                    label_cell.fill = PatternFill(
                        start_color='F3F4F6',
                        end_color='F3F4F6',
                        fill_type='solid'
                    )
                    
                    value_cell = ws_summary.cell(row=row_idx, column=2, value=value)
                    value_cell.font = Font(name='Arial', size=10)
                
                # 报告总结
                if report_data.summary:
                    summary_title_cell = ws_summary.cell(
                        row=len(info_data) + 4,
                        column=1,
                        value="Summary Text"
                    )
                    summary_title_cell.font = Font(name='Arial', size=11, bold=True)
                    
                    summary_cell = ws_summary.cell(
                        row=len(info_data) + 5,
                        column=1,
                        value=report_data.summary
                    )
                    summary_cell.font = Font(name='Arial', size=10)
                    summary_cell.alignment = Alignment(
                        horizontal='left',
                        vertical='top',
                        wrap_text=True
                    )
                    ws_summary.merge_cells(
                        start_row=len(info_data) + 5,
                        start_column=1,
                        end_row=len(info_data) + 5,
                        end_column=2
                    )
                
                # 调整列宽
                ws_summary.column_dimensions['A'].width = 20
                ws_summary.column_dimensions['B'].width = 60
                
                logger.debug("总结工作表已创建")
            
            # 工作表3: 元信息
            if include_metadata:
                ws_meta = wb.create_sheet("Metadata", 2)
                
                # 标题
                title_cell = ws_meta.cell(row=1, column=1, value="Column Metadata")
                title_cell.font = Font(name='Arial', size=14, bold=True, color='1F2937')
                ws_meta.merge_cells('A1:C1')
                
                # 表头
                meta_headers = ["Column Name", "Data Type", "Sample Values"]
                for col_idx, header in enumerate(meta_headers, 1):
                    cell = ws_meta.cell(row=3, column=col_idx, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = border
                
                # 列信息
                for row_idx, col_name in enumerate(headers, 4):
                    # 列名
                    name_cell = ws_meta.cell(row=row_idx, column=1, value=col_name)
                    name_cell.font = data_font
                    name_cell.border = border
                    
                    # 数据类型
                    col_type = report_data.metadata.column_types.get(col_name, 'unknown')
                    type_cell = ws_meta.cell(row=row_idx, column=2, value=col_type)
                    type_cell.font = data_font
                    type_cell.border = border
                    
                    # 示例值（前3个非空值）
                    sample_values = []
                    for row_data in report_data.data[:10]:
                        value = row_data.get(col_name)
                        if value is not None and value != '':
                            sample_values.append(str(value))
                            if len(sample_values) >= 3:
                                break
                    
                    sample_text = ', '.join(sample_values) if sample_values else 'N/A'
                    sample_cell = ws_meta.cell(row=row_idx, column=3, value=sample_text)
                    sample_cell.font = data_font
                    sample_cell.border = border
                    
                    # 交替行背景色
                    if row_idx % 2 == 0:
                        for col_idx in range(1, 4):
                            ws_meta.cell(row=row_idx, column=col_idx).fill = PatternFill(
                                start_color='F9FAFB',
                                end_color='F9FAFB',
                                fill_type='solid'
                            )
                
                # 调整列宽
                ws_meta.column_dimensions['A'].width = 25
                ws_meta.column_dimensions['B'].width = 15
                ws_meta.column_dimensions['C'].width = 40
                
                logger.debug("元信息工作表已创建")
            
            # 保存到字节流
            buffer = BytesIO()
            wb.save(buffer)
            excel_bytes = buffer.getvalue()
            buffer.close()
            
            logger.info(
                f"Excel生成完成: size={len(excel_bytes)} bytes, "
                f"sheets={len(wb.sheetnames)}"
            )
            
            return excel_bytes
            
        except Exception as e:
            logger.error(
                f"Excel生成失败",
                extra={
                    "title": report_data.title,
                    "data_rows": len(report_data.data),
                    "error": str(e)
                },
                exc_info=True
            )
            raise Exception(f"Excel生成失败: {str(e)}")


# 全局导出服务实例
_export_service = None


def get_export_service() -> ExportService:
    """
    获取全局导出服务实例
    
    Returns:
        ExportService实例
    """
    global _export_service
    
    if _export_service is None:
        _export_service = ExportService()
    
    return _export_service
